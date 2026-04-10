from __future__ import annotations

import io
from datetime import datetime, timezone


def build_conversation_context(history: list[dict[str, object]], *, max_turns: int = 20) -> list[dict[str, str]]:
    """Builds bounded user/assistant turn context for router requests."""
    context: list[dict[str, str]] = []
    for message in history:
        role = str(message.get("role", "")).strip().lower()
        if role not in {"user", "assistant"}:
            continue
        if role == "assistant" and message.get("answer"):
            content = str(message.get("answer", "")).strip()
        else:
            content = str(message.get("content", "")).strip()
        if not content:
            continue
        context.append({"role": role, "content": content})
    return context[-max_turns:]


def chat_history_markdown(history: list[dict[str, object]]) -> str:
    """Renders chat history as a markdown transcript with metadata and citations."""
    header = f"# Chat Export\n\nGenerated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}\n\n"
    if not history:
        return header + "_No chat messages yet._\n"

    lines: list[str] = [header]
    for idx, msg in enumerate(history, start=1):
        role = str(msg.get("role", "unknown")).strip().title()
        if role == "Assistant" and msg.get("answer"):
            body = str(msg.get("answer", "")).strip()
        else:
            body = str(msg.get("content", "")).strip()
        if not body:
            continue
        lines.append(f"## {idx}. {role}\n")
        lines.append(f"{body}\n")

        meta = str(msg.get("meta", "")).strip()
        if meta:
            lines.append(f"- Meta: `{meta}`\n")
        sources = msg.get("sources", [])
        if isinstance(sources, list) and sources:
            lines.append("- Sources:\n")
            for source in sources:
                if not isinstance(source, dict):
                    continue
                marker = str(source.get("marker", ""))
                label = str(source.get("label", ""))
                if marker or label:
                    lines.append(f"  - {marker} {label}".rstrip() + "\n")
        lines.append("\n")
    return "".join(lines)


def citation_markers(sources: list[dict[str, object]], *, max_markers: int = 4) -> str:
    """Returns compact citation markers string like `[1] [2]`."""
    if not sources:
        return ""
    markers: list[str] = []
    for source in sources[:max_markers]:
        marker = str(source.get("marker", "")).strip()
        if marker:
            markers.append(marker)
    return " ".join(markers)


def export_chat_xlsx(history: list[dict[str, object]]) -> bytes:
    """Build an Excel workbook from chat history and return raw bytes.

    Each assistant turn becomes one row with columns: Question, Answer,
    Sources, Confidence, Answer Mode, Trace ID, Timestamp.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chat Export"

    headers = ["#", "Question", "Answer", "Sources", "Confidence", "Answer Mode", "Trace ID"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    row_num = 1
    for msg in history:
        if str(msg.get("role", "")) != "assistant":
            continue
        if not msg.get("answer"):
            continue
        row_num += 1
        sources = msg.get("sources", [])
        source_text = "; ".join(
            str(s.get("label", "")) for s in (sources if isinstance(sources, list) else []) if s.get("label")
        )
        confidence = msg.get("confidence_score")
        conf_text = f"{float(confidence):.0%}" if isinstance(confidence, (int, float)) else "N/A"
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=str(msg.get("question", ""))).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=3, value=str(msg.get("answer", ""))).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=4, value=source_text)
        ws.cell(row=row_num, column=5, value=conf_text)
        ws.cell(row=row_num, column=6, value=str(msg.get("answer_mode", "")))
        ws.cell(row=row_num, column=7, value=str(msg.get("trace_id", "")))

    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_smart_actions(
    answer_mode: str,
    sources: list[dict[str, object]],
    confidence_score: float | None,
    sticky_survey: str | None,
) -> list[str]:
    """Generate contextual next-action suggestion labels based on the current response.

    The suggestions are rendered as clickable chips in the UI.  Clicking one
    pre-fills the chat input with the action text so the user can send it
    directly or edit it first.

    Parameters
    ----------
    answer_mode : str
        The ``answer_mode`` returned by the router (e.g. ``direct_answer``,
        ``allowed_values``, ``summary``, ``clarifier``, ``lineage``).
    sources : list[dict[str, object]]
        Citation source dicts from the router response.  Used to detect when
        results span multiple surveys.
    confidence_score : float | None
        Retrieval confidence score (0–1).  Low scores surface rephrase hints.
    sticky_survey : str | None
        The currently active survey context, if any.

    Returns
    -------
    list[str]
        Up to 4 short action label strings.
    """
    actions: list[str] = []

    # Collect distinct survey names from sources
    survey_names: list[str] = []
    for src in (sources or []):
        name = str(src.get("survey_name", "") or src.get("label", "")).strip()
        if name and name not in survey_names:
            survey_names.append(name)

    # Answer-mode specific suggestions
    if answer_mode == "allowed_values":
        actions.append("Compare response scales across waves")
        actions.append("Show all questions in this section")
    elif answer_mode == "summary":
        actions.append("Break down by individual question")
        actions.append("Compare across years")
    elif answer_mode == "clarifier":
        actions.append("Show all variants")
        actions.append("Filter to one survey")
    elif answer_mode == "direct_answer":
        actions.append("Find broader concepts")
        actions.append("Show similar wording")
    elif answer_mode == "lineage":
        actions.append("Compare across surveys")
        actions.append("Show allowed values")

    # Multi-survey: offer focused filter chips
    if len(survey_names) >= 2:
        actions.insert(0, f"Filter to {survey_names[0]} only")

    # Low confidence: suggest rephrasing
    if isinstance(confidence_score, float) and confidence_score < 0.35:
        actions.append("Rephrase my question")

    # Always offer lineage unless we're already in lineage mode
    if answer_mode != "lineage":
        actions.append("View lineage across waves")

    # Deduplicate while preserving order, cap at 4
    seen: set[str] = set()
    unique: list[str] = []
    for action in actions:
        if action not in seen:
            seen.add(action)
            unique.append(action)
        if len(unique) == 4:
            break
    return unique


def confidence_badge(confidence_score: float | None) -> str:
    """Returns user-facing confidence label with percentage and plain-language description."""
    if confidence_score is None:
        return "Match quality: exact lookup"
    pct = f"{confidence_score:.0%}"
    if confidence_score >= 0.75:
        return f"Match quality: High ({pct}) — strong retrieval match"
    if confidence_score >= 0.45:
        return f"Match quality: Medium ({pct}) — good retrieval match"
    if confidence_score >= 0.25:
        return f"Match quality: Low ({pct}) — partial match, review results"
    return f"Match quality: Very low ({pct}) — consider rephrasing your query"
